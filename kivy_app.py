import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout

EXCEL_FILE = 'Arihant_Agro.xlsx'

# Initialize Excel file with column names and data types
def initialize_excel():
    workbook = Workbook()
    sheet_names = ['Customers', 'Products', 'Purchases', 'Sales', 'Payments']
    columns = {
        'Customers': ['Name', 'Contact', 'Address'],
        'Products': ['Name', 'Quantity', 'Price', 'Unit'],
        'Purchases': ['Date', 'Invoice', 'Product', 'Quantity', 'Customer', 'Total Price', 'GST'],
        'Sales': ['Date', 'Total Price', 'Product', 'Customer', 'Challan', 'Quantity', 'GST'],
        'Payments': ['Invoice', 'Date', 'Amount', 'Method', 'Note', 'Customer', 'Direction']
    }
    
    for sheet_name in sheet_names:
        sheet = workbook.create_sheet(sheet_name)
        for idx, column_name in enumerate(columns[sheet_name], start=1):
            cell = sheet.cell(row=1, column=idx)
            cell.value = column_name

    workbook.save(EXCEL_FILE)

# Ensure the sheet exists in the Excel file
def create_sheet_if_not_exists(sheet_name):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    workbook.save(EXCEL_FILE)

# Append a row of data to the specified sheet
def append_data_to_sheet(sheet_name, data):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    sheet.append(data)
    workbook.save(EXCEL_FILE)

# Get all rows of data from the specified sheet
def get_all_data_from_sheet(sheet_name):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    return list(sheet.iter_rows(values_only=True))

# Update a specific row of data in the specified sheet
def update_row_in_sheet(sheet_name, row_index, data):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    for idx, value in enumerate(data, start=1):
        sheet.cell(row=row_index, column=idx).value = value
    workbook.save(EXCEL_FILE)

# Delete a specific row from the specified sheet
def delete_row_in_sheet(sheet_name, row_index):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    sheet.delete_rows(row_index, 1)
    workbook.save(EXCEL_FILE)

# Customer Management Functions
def create_customer(name, contact, address):
    create_sheet_if_not_exists('Customers')
    append_data_to_sheet('Customers', [name, contact, address])

def display_customers():
    create_sheet_if_not_exists('Customers')
    return get_all_data_from_sheet('Customers')[1:]  # Skip header row

def update_customer(row_index, name, contact, address):
    update_row_in_sheet('Customers', row_index, [name, contact, address])

def delete_customer(row_index):
    delete_row_in_sheet('Customers', row_index)

# Product Management Functions
def create_product(name, quantity, price, unit):
    create_sheet_if_not_exists('Products')
    append_data_to_sheet('Products', [name, quantity, price, unit])

def display_products():
    create_sheet_if_not_exists('Products')
    return get_all_data_from_sheet('Products')[1:]  # Skip header row

def update_product(row_index, name, quantity, price, unit):
    update_row_in_sheet('Products', row_index, [name, quantity, price, unit])

def delete_product(row_index):
    delete_row_in_sheet('Products', row_index)

# Purchase Management Functions
def create_purchase(date, invoice, product, quantity, customer, total_price, gst):
    create_sheet_if_not_exists('Purchases')
    append_data_to_sheet('Purchases', [date, invoice, product, quantity, customer, total_price, gst])

def display_purchases():
    create_sheet_if_not_exists('Purchases')
    return get_all_data_from_sheet('Purchases')[1:]  # Skip header row

def delete_purchase(row_index):
    delete_row_in_sheet('Purchases', row_index)

# Sale Management Functions
def create_sale(date, total_price, product, customer, challan, quantity, gst):
    create_sheet_if_not_exists('Sales')
    append_data_to_sheet('Sales', [date, total_price, product, customer, challan, quantity, gst])

def display_sales():
    create_sheet_if_not_exists('Sales')
    return get_all_data_from_sheet('Sales')[1:]  # Skip header row

def delete_sale(row_index):
    delete_row_in_sheet('Sales', row_index)

# Payment Management Functions
def create_payment(invoice, date, amount, method, note, customer, direction):
    create_sheet_if_not_exists('Payments')
    append_data_to_sheet('Payments', [invoice, date, amount, method, note, customer, direction])

def display_payments():
    create_sheet_if_not_exists('Payments')
    return get_all_data_from_sheet('Payments')[1:]  # Skip header row

def delete_payment(row_index):
    delete_row_in_sheet('Payments', row_index)

# Kivy App Interface
class ArihantAgroApp(App):
    def build(self):
        self.title = 'Arihant Agro Management System'
        self.main_layout = BoxLayout(orientation='vertical')

        # Customer Management
        self.customer_layout = BoxLayout(orientation='horizontal')
        self.customer_name_input = TextInput(hint_text='Customer Name')
        self.customer_contact_input = TextInput(hint_text='Contact')
        self.customer_address_input = TextInput(hint_text='Address')
        self.add_customer_button = Button(text='Add Customer', on_press=self.add_customer)
        self.show_customers_button = Button(text='Show Customers', on_press=self.show_customers)
        self.customer_layout.add_widget(self.customer_name_input)
        self.customer_layout.add_widget(self.customer_contact_input)
        self.customer_layout.add_widget(self.customer_address_input)
        self.customer_layout.add_widget(self.add_customer_button)
        self.main_layout.add_widget(self.customer_layout)
        self.main_layout.add_widget(self.show_customers_button)

        # Product Management
        self.product_layout = BoxLayout(orientation='horizontal')
        self.product_name_input = TextInput(hint_text='Product Name')
        self.product_quantity_input = TextInput(hint_text='Quantity')
        self.product_price_input = TextInput(hint_text='Price')
        self.product_unit_input = TextInput(hint_text='Unit')
        self.add_product_button = Button(text='Add Product', on_press=self.add_product)
        self.show_products_button = Button(text='Show Products', on_press=self.show_products)
        self.product_layout.add_widget(self.product_name_input)
        self.product_layout.add_widget(self.product_quantity_input)
        self.product_layout.add_widget(self.product_price_input)
        self.product_layout.add_widget(self.product_unit_input)
        self.product_layout.add_widget(self.add_product_button)
        self.main_layout.add_widget(self.product_layout)
        self.main_layout.add_widget(self.show_products_button)

        # Purchase Management
        self.purchase_layout = BoxLayout(orientation='horizontal')
        self.purchase_date_input = TextInput(hint_text='Purchase Date (YYYY-MM-DD)')
        self.invoice_number_input = TextInput(hint_text='Invoice Number')
        self.purchase_product_input = TextInput(hint_text='Product')
        self.purchase_quantity_input = TextInput(hint_text='Quantity')
        self.purchase_customer_input = TextInput(hint_text='Customer')
        self.purchase_total_price_input = TextInput(hint_text='Total Price')
        self.purchase_gst_input = TextInput(hint_text='GST')
        self.add_purchase_button = Button(text='Add Purchase', on_press=self.add_purchase)
        self.show_purchases_button = Button(text='Show Purchases', on_press=self.show_purchases)
        self.purchase_layout.add_widget(self.purchase_date_input)
        self.purchase_layout.add_widget(self.invoice_number_input)
        self.purchase_layout.add_widget(self.purchase_product_input)
        self.purchase_layout.add_widget(self.purchase_quantity_input)
        self.purchase_layout.add_widget(self.purchase_customer_input)
        self.purchase_layout.add_widget(self.purchase_total_price_input)
        self.purchase_layout.add_widget(self.purchase_gst_input)
        self.purchase_layout.add_widget(self.add_purchase_button)
        self.main_layout.add_widget(self.purchase_layout)
        self.main_layout.add_widget(self.show_purchases_button)

        # Sale Management
        self.sale_layout = BoxLayout(orientation='horizontal')
        self.sale_date_input = TextInput(hint_text='Sale Date (YYYY-MM-DD)')
        self.sale_total_price_input = TextInput(hint_text='Total Price')
        self.sale_product_name_input = TextInput(hint_text='Product Name')
        self.sale_customer_name_input = TextInput(hint_text='Customer Name')
        self.challan_number_input = TextInput(hint_text='Challan Number')
        self.sale_quantity_input = TextInput(hint_text='Quantity')
        self.sale_gst_input = TextInput(hint_text='GST')
        self.add_sale_button = Button(text='Add Sale', on_press=self.add_sale)
        self.show_sales_button = Button(text='Show Sales', on_press=self.show_sales)
        self.sale_layout.add_widget(self.sale_date_input)
        self.sale_layout.add_widget(self.sale_total_price_input)
        self.sale_layout.add_widget(self.sale_product_name_input)
        self.sale_layout.add_widget(self.sale_customer_name_input)
        self.sale_layout.add_widget(self.challan_number_input)
        self.sale_layout.add_widget(self.sale_quantity_input)
        self.sale_layout.add_widget(self.sale_gst_input)
        self.sale_layout.add_widget(self.add_sale_button)
        self.main_layout.add_widget(self.sale_layout)
        self.main_layout.add_widget(self.show_sales_button)

        # Payment Management
        self.payment_layout = BoxLayout(orientation='horizontal')
        self.payment_invoice_input = TextInput(hint_text='Invoice')
        self.payment_date_input = TextInput(hint_text='Date (YYYY-MM-DD)')
        self.payment_amount_input = TextInput(hint_text='Amount')
        self.payment_method_input = TextInput(hint_text='Method')
        self.payment_note_input = TextInput(hint_text='Note')
        self.payment_customer_input = TextInput(hint_text='Customer')
        self.payment_direction_input = TextInput(hint_text='Direction')
        self.add_payment_button = Button(text='Add Payment', on_press=self.add_payment)
        self.show_payments_button = Button(text='Show Payments', on_press=self.show_payments)
        self.payment_layout.add_widget(self.payment_invoice_input)
        self.payment_layout.add_widget(self.payment_date_input)
        self.payment_layout.add_widget(self.payment_amount_input)
        self.payment_layout.add_widget(self.payment_method_input)
        self.payment_layout.add_widget(self.payment_note_input)
        self.payment_layout.add_widget(self.payment_customer_input)
        self.payment_layout.add_widget(self.payment_direction_input)
        self.payment_layout.add_widget(self.add_payment_button)
        self.main_layout.add_widget(self.payment_layout)
        self.main_layout.add_widget(self.show_payments_button)

        return self.main_layout

    # Customer Management Handlers
    def add_customer(self, instance):
        name = self.customer_name_input.text
        contact = self.customer_contact_input.text
        address = self.customer_address_input.text
        create_customer(name, contact, address)
        self.show_popup('Customer Added', f'Customer {name} has been added successfully.')

    def show_customers(self, instance):
        customers = display_customers()
        self.show_data_popup('Customers', customers)

    # Product Management Handlers
    def add_product(self, instance):
        name = self.product_name_input.text
        quantity = self.product_quantity_input.text
        price = self.product_price_input.text
        unit = self.product_unit_input.text
        create_product(name, quantity, price, unit)
        self.show_popup('Product Added', f'Product {name} has been added successfully.')

    def show_products(self, instance):
        products = display_products()
        self.show_data_popup('Products', products)

    # Purchase Management Handlers
    def add_purchase(self, instance):
        date = self.purchase_date_input.text
        invoice = self.invoice_number_input.text
        product = self.purchase_product_input.text
        quantity = self.purchase_quantity_input.text
        customer = self.purchase_customer_input.text
        total_price = self.purchase_total_price_input.text
        gst = self.purchase_gst_input.text
        create_purchase(date, invoice, product, quantity, customer, total_price, gst)
        self.show_popup('Purchase Added', f'Purchase {invoice} has been added successfully.')

    def show_purchases(self, instance):
        purchases = display_purchases()
        self.show_data_popup('Purchases', purchases)

    # Sale Management Handlers
    def add_sale(self, instance):
        date = self.sale_date_input.text
        total_price = self.sale_total_price_input.text
        product = self.sale_product_name_input.text
        customer = self.sale_customer_name_input.text
        challan = self.challan_number_input.text
        quantity = self.sale_quantity_input.text
        gst = self.sale_gst_input.text
        create_sale(date, total_price, product, customer, challan, quantity, gst)
        self.show_popup('Sale Added', f'Sale {challan} has been added successfully.')

    def show_sales(self, instance):
        sales = display_sales()
        self.show_data_popup('Sales', sales)

    # Payment Management Handlers
    def add_payment(self, instance):
        invoice = self.payment_invoice_input.text
        date = self.payment_date_input.text
        amount = self.payment_amount_input.text
        method = self.payment_method_input.text
        note = self.payment_note_input.text
        customer = self.payment_customer_input.text
        direction = self.payment_direction_input.text
        create_payment(invoice, date, amount, method, note, customer, direction)
        self.show_popup('Payment Added', f'Payment {invoice} has been added successfully.')

    def show_payments(self, instance):
        payments = display_payments()
        self.show_data_popup('Payments', payments)

    # Utility Methods
    def show_popup(self, title, message):
        popup_layout = BoxLayout(orientation='vertical')
        popup_label = Label(text=message)
        close_button = Button(text='Close', on_press=lambda *args: popup.dismiss())
        popup_layout.add_widget(popup_label)
        popup_layout.add_widget(close_button)
        popup = Popup(title=title, content=popup_layout, size_hint=(0.5, 0.5))
        popup.open()

    def show_data_popup(self, title, data):
        popup_layout = BoxLayout(orientation='vertical')
        scroll_view = ScrollView()
        grid_layout = GridLayout(cols=len(data[0]), size_hint_y=None)
        grid_layout.bind(minimum_height=grid_layout.setter('height'))

        for row in data:
            for item in row:
                grid_layout.add_widget(Label(text=str(item)))

        scroll_view.add_widget(grid_layout)
        close_button = Button(text='Close', on_press=lambda *args: popup.dismiss())
        popup_layout.add_widget(scroll_view)
        popup_layout.add_widget(close_button)
        popup = Popup(title=title, content=popup_layout, size_hint=(0.9, 0.9))
        popup.open()

if __name__ == '__main__':
    initialize_excel()
    ArihantAgroApp().run()
